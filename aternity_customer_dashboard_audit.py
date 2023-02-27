
import csv
import argparse
import getpass
import json
import os
from datetime import datetime,timedelta
import calendar
import pprint
import time
import logging
import shutil
from pathlib import Path

#import openpyxl
#import requests

try:
    import openpyxl
except ImportError:
    import pip
    pip.main(['install', '--user', 'openpyxl'])
    import openpyxl

try:
    import requests
except ImportError:
    import pip
    pip.main(['install', '--user', 'requests'])
    import requests

__author__ = 'Nick Hurst'
__author_email__ = 'nick.hurst@riverbed.com'
__version__ = '1.0.0'


DEFAULT_CUSTOMER_ATERNITY_INSTANCE_FILE_PATH = "account_info.csv"
DEBUG_LEVEL_LIST = ('10','20','30','40','50')

class CustomerAternityInstance:

    def __init__(self,customer_name:str,aternity_web_instance_url:str,number_of_days :int, username:str, passwd:str) -> None:

        self.audit_report_work_book = openpyxl.Workbook()
        
        #std=self.audit_report_work_book('Sheet')
        
        logging.debug(f'audit_report_work_book created')
        logging.debug(f'')

        self.customer_name = customer_name
        logging.debug(f'class instance create for {self.customer_name}')
        #convert standard web url to odat/restapi url
        stage1 = aternity_web_instance_url.split('.',1)
        stage2 = stage1[0].split('/',-1)
        stage3 = f"{stage2[2]}-odata"
        stage4 = f"https://{stage3}.{stage1[1]}"
        self.aternity_instance_url = stage4

        self.number_of_days = number_of_days
        self.username = username
        self.passwd = passwd

        #if "data" folder doesn't exist, then create. Folder required for location to store results
        if not os.path.exists('./data'):
            logging.warning("data folder not present - creating")
            os.mkdir("data")

        self.week_day_data_structure = self.create_report_week_days_data_structure()
        #for loading data for testing reports, should normally be marked out   
        #self.load_old_reports()
        
    def load_old_reports(self):
        #this function loads previous saved rest api responses, for use in testing reports

        #check if "data" folder, in directory that script is run from, so see if there is a aternity dashboard audit file for this customer
        list_of_files_in_folder =  os.listdir("data/")
        for file_name in list_of_files_in_folder:
            #print(self.customer_name)
            if file_name.__contains__("Aternity_dash_audit") and file_name.__contains__(self.customer_name):
                print(file_name)
                with open(f"data/{file_name}") as file_data:
                    self.previous_responces = json.load(file_data)
        try:
            self.report_repsonse = self.previous_responces
        except:
            pass
        self.parse_response_into_day_week_data_structure()

    def find_data_on_date(self,data_date:str):
        logging.debug(f'method - find_data_on_date')
        user_list = []
        dashboard_list = []
        for value in self.report_repsonse['value']:
            split_timestamp = value['AUDIT_ACTION_TIMESTAMP'].split('T')
            if split_timestamp[0] == data_date:
                user_list.append(value['USER_NAME'])
                dashboard_list.append(value['DASHBOARD_NAME'])
        return user_list,dashboard_list

    def parse_response_into_day_week_data_structure(self)->None:
        logging.debug(f'method - parse_response_into_day_week_data_structure')
        for week in range(0,len(self.week_day_data_structure)):
            for day in self.week_day_data_structure[week]:
                day['users'],day['dashboards'] = self.find_data_on_date(day['date'])
        #with open(f"data/{self.customer_name}_Weekly daily.json","w") as save_data:
            #json.dump(self.week_day_data_structure,save_data,indent=6)

    def run_rest_api_report_request(self,url_default_override :str = "",flag_called_by_self :bool = False)-> bool:
        logging.debug(f'method - run_rest_api_report_request')
        #rest API call to aternity server to get user dashboard usage report
        """
        flag "flag_called_by_self" to change return. instead of boolean to list.
        reason for this, recursion.
        If the rest api request produces to much data, then the results are split and spread over mulitple requests,
        at the end of the first response (and proceding responses) is an url under key "@odata.nextLink", for the
        next part of the results. Once there is a response without the key "@odata.nextLink", then we know the 
        report is finished. When we key "@odata.nextLink" we do recursion, calling the the function again, however
        with "flag_called_by_self" set to True and "url_default_override" set to the url found in "@odata.nextLink",
        with each call returning its results.
        """
        flag_report_successful = False
        default_url = f"{self.aternity_instance_url}aternity.odata/latest/DASHBOARD_VIEWS_AUDIT?$filter=relative_time(last_{str(self.number_of_days)}_days) and endswith(account_name,'{self.customer_name}')"
        logging.info(f'api url - {default_url}')
        #create url of rest API with customer instance address, number of days to report on and customer name
        if not url_default_override:
            url = default_url
        else:
            url = url_default_override

            
        headersList = {
            "Accept": "*/*"
        }

        payload = ""
        try:
            response = requests.request("GET", url, data=payload,  headers=headersList,verify=False, auth=(self.username,self.passwd))
            logging.info(f'api response{response}')
            logging.debug(f'api text response{response.text}')
            result = response.json()            
        except:
            flag_report_successful = False
            logging.warning(f'request failed due to - {response} - {response.text}')
        else:
            result = response.json()
            logging.debug(f'returned data type is - {type(result)}')
            if type(result) is dict:

                self.report_repsonse = result
                flag_report_successful = True

                if result.get("@odata.nextLink"):
                    #if old response has data
                    logging.info(f'result has odata.nextLink - {result.get("@odata.nextLink")}')
                    self.report_repsonse["value"] = self.report_repsonse["value"] + self.run_rest_api_report_request(self.report_repsonse["@odata.nextLink"],True)["value"]
                                        
                #check if response doesn't have any data in "value" key of response. This can happen with you authenticated but not authorized to access the customer artnity instance
                #if no data is present then we set flag_report_successful to False
                if not self.report_repsonse['value']:
                    logging.warning(f'value key not present {result}')
                    flag_report_successful = False
                else:
                    #save response to local machine in "data" folder
                    flag_report_successful = True
                    with open(f"data/Aternity_RAW_report_{self.customer_name}.json","w")as save_file:
                        logging.info(f'save results to file "data/Aternity_RAW_report_{self.customer_name}.json')
                        json.dump(self.report_repsonse,save_file,indent=6)
            else:
                logging.warning(f'returned data type is - {type(result)}')
                logging.warning(f'report failed due to - {response} - {result}')
                flag_report_successful = False

        
        if flag_called_by_self:
            return self.report_repsonse
        else:
            if flag_report_successful: 
                self.parse_response_into_day_week_data_structure()
            else: 
                print(f'unable to run report for {self.customer_name}')
                logging.warning(f'unable to run report for {self.customer_name}')
            return flag_report_successful

    def create_report_week_days_data_structure(self)-> dict:
        logging.debug(f'method - create_report_week_days_data_structure')
        #get todays date, then work out date number of days user has request to filter by
        last_day_of_report = datetime.now()
        d = timedelta(days = self.number_of_days)
        filter_start_date = last_day_of_report - d

        week_number = 0
        all_weeks = {}
        day_index = 180
        #loop while day index is less tht date of last whole week
        day_index = filter_start_date
        day_count = (last_day_of_report - day_index).days
        week_data=[]
        while day_count >=  0:
            
            
            day_index_num =  (calendar.weekday(day_index.year, day_index.month, day_index.day))
            #print(f"{day_count}\t{day_index}\t day num {day_index_num}")

            day_data = {}
            #insure date is same formate as date format in report from rest api call, as no single digitis 1 should be 01
            if (day_index.month in range(0,10)) and (day_index.day in range(0,10)):
                day_data['date'] = f'{day_index.year}-0{day_index.month}-0{day_index.day}'
            elif day_index.day in range(0,10):
                day_data['date'] = f'{day_index.year}-{day_index.month}-0{day_index.day}'
            elif day_index.month in range(0,10):
                day_data['date'] = f'{day_index.year}-0{day_index.month}-{day_index.day}'
            else:
                day_data['date'] = f'{day_index.year}-{day_index.month}-{day_index.day}'

            day_data['users'] = []
            day_data['dashboards'] = []
            day_data['day_number'] = day_index_num
            week_data.append(day_data)

            day_index = day_index + timedelta(1)

            if day_index_num == 6:
                all_weeks[week_number]=week_data
                week_data = []
                week_number +=1
            elif day_count == 0:
                all_weeks[week_number]=week_data
                week_data = []
                week_number +=1
            day_count = (last_day_of_report - day_index).days

        return all_weeks

    def get_list_of_user_using_aternity(self) -> list[str]:
        logging.debug(f'method - get_list_of_user_using_aternity')
        #go through rest api response, get list of all users that have used the customer aternity instance and return list
        list_of_users = []
        for value in self.report_repsonse["value"]:
            if not value['USER_NAME'] in list_of_users:
                list_of_users.append(value['USER_NAME'])
        return list_of_users
    
    def get_list_of_dashboard_used(self) -> list[str]:
        logging.debug(f'method - get_list_of_dashboard_used')
        #go through rest api response, get list of all dashboards used the in customer aternity instance and return list
        list_of_used_dashboards = []
        for value in self.report_repsonse["value"]:
            if not value['DASHBOARD_NAME'] in list_of_used_dashboards:
                list_of_used_dashboards.append(value['DASHBOARD_NAME'])
        return list_of_used_dashboards

    def what_dashboard_did_user_use(self,username)-> list[str]:
        logging.debug(f'method - what_dashboard_did_user_use')
        #go through rest api response, with username provided, create list of all dashboards that the user has used in customer aternity instance and return list
        dashboard_used_list = []
        for value in self.report_repsonse['value']:
            if value['USER_NAME'] == username:
                if not value['DASHBOARD_NAME'] in dashboard_used_list:
                    dashboard_used_list.append(value['DASHBOARD_NAME'])
        return dashboard_used_list
    
    def what_users_used_dashboard(self,dashboard_name) -> list[str]:
        logging.debug(f'method - what_users_used_dashboard')
        #go through rest api response, with dashboard name provided, create list of all users that have used the dashbaord in customer aternity instance and return list    
        users_that_used_dashboard_list = []
        for value in self.report_repsonse['value']:
            if value['DASHBOARD_NAME'] == dashboard_name:
                if not value['USER_NAME'] in users_that_used_dashboard_list:
                    users_that_used_dashboard_list.append(value['USER_NAME'])
        return users_that_used_dashboard_list

    def report_of_users_using_aternity(self) -> None:
        logging.debug(f'method - report_of_users_using_aternity')
        #called report, result being a text file saved to "data" folder, listing all users that have used the customer aternity instance
        with open(f"data/{self.customer_name}_List of users using aternity.txt","w") as save_data:
            for row in self.get_list_of_user_using_aternity():
                save_data.write(f'{row}\n')

    def report_what_dashboard_users_used(self)-> None:
        logging.debug(f'method - report_what_dashboard_users_used')
        #called report, result being a json file saved to "data" folder, listing all users that have used the customer aternity instance and which dashboards they used
        user_used_dashboard_list = []
        for user in self.get_list_of_user_using_aternity():
            data={}
            data['username']=user
            data['dashboards_used'] = self.what_dashboard_did_user_use(user)
            user_used_dashboard_list.append(data)
        with open(f"data/{self.customer_name}_List of dashboards users used in aternity.json","w") as save_data:
            json.dump(user_used_dashboard_list,save_data,indent=6)
    
    def report_what_user_used_which_dashboard(self)-> None:
        logging.debug(f'method - report_what_user_used_which_dashboard')
        #called report, result being a json file saved to "data" folder, listing all dashboard accessed in the customer aternity instance and which users used them
        dashboards_used_by_users_list = []
        for dashboard in self.get_list_of_dashboard_used():
            data={}
            data['dashboard_name'] = dashboard
            data['users_that_used_dashboard'] = self.what_users_used_dashboard(dashboard)
            dashboards_used_by_users_list.append(data)
        with open(f"data/{self.customer_name}_List of dashboards and what users used them in aternity.json","w") as save_data:
            json.dump(dashboards_used_by_users_list,save_data,indent=6)

    def report_excel_spreadsheet_user_and_dashboard_overview(self)->None:
        logging.debug(f'method - report_excel_spreadsheet_user_and_dashboard_overview')
        row_start = 2
        col_start = 2


        self.audit_report_work_book.create_sheet("User_and_Dashboard_Matrix")
        user_and_dashboard_work_sheet = self.audit_report_work_book["User_and_Dashboard_Matrix"]

        #print dashboard names in top row
        for col, dash_name in enumerate(self.get_list_of_dashboard_used()):
            user_and_dashboard_work_sheet.cell(1,col+col_start).value = dash_name
        
        #print usernames in first column
        for row, username in enumerate(self.get_list_of_user_using_aternity()):
            user_and_dashboard_work_sheet.cell(row+row_start,1).value = username

        #add "x" mark to cell showing what dashboards users have used
        for row,username in enumerate(self.get_list_of_user_using_aternity()):
            for dash_viewed_name in self.what_dashboard_did_user_use(username):
                for col,bash_name in enumerate(self.get_list_of_dashboard_used()):
                    if dash_viewed_name == bash_name:
                        user_and_dashboard_work_sheet.cell(row+row_start,col+col_start).value = "x"

    def report_excel_spreadsheet_user_daily_dashboard_usage(self)->None:
        logging.debug(f'method - report_excel_spreadsheet_user_daily_dashboard_usage')
        """
        report show the number total dash boards a user has viewed each day
        """

        #setup worksheet
        self.audit_report_work_book.create_sheet("user_and_daily_dashboard_usage")
        user_daily_dashboard_work_sheet = self.audit_report_work_book["user_and_daily_dashboard_usage"]
        
        #create column tittles
        user_daily_dashboard_work_sheet.cell(1,1).value = "week number"
        user_daily_dashboard_work_sheet.cell(1,2).value = "date"
        user_daily_dashboard_work_sheet.cell(1,3).value = "day"
        name_col = 4
        for user in self.get_list_of_user_using_aternity():
            #print user names in column tittles
            user_daily_dashboard_work_sheet.cell(1,name_col).value = user
            name_col+=1


         #output data to worksheet
        row_number = 2
        merged_cells = True
        for w_data in self.week_day_data_structure:
            for d_data in self.week_day_data_structure[w_data]:

                #for week number we those cells will be mergered for each week
                if merged_cells:
                    user_daily_dashboard_work_sheet.merge_cells(start_row=row_number, start_column=1,end_row=(row_number+len(self.week_day_data_structure[w_data])-1), end_column=1)
                    user_daily_dashboard_work_sheet.cell(row_number,1).value = w_data
                    merged_cells = False
                if d_data['day_number'] == 6:
                    merged_cells = True
                    
                user_daily_dashboard_work_sheet.cell(row_number,2).value = d_data['date']
                user_daily_dashboard_work_sheet.cell(row_number,3).value = self.day_name(d_data['day_number'])
                name_col = 4
                for user in self.get_list_of_user_using_aternity():
                    user_daily_dashboard_work_sheet.cell(row_number,name_col).value = d_data['users'].count(user)
                    name_col+=1

                row_number += 1

    def excel_audit_report(self):
        logging.debug(f'method - excel_audit_report  for {self.customer_name}')
        self.report_excel_spreadsheet_user_and_dashboard_overview()
        self.report_excel_spreadsheet_user_daily_dashboard_usage()
        self.report_excel_understand_daily_usage()

        
        #remove default first sheet created when wookbook was created, then save workbook
        try:
            del self.audit_report_work_book['Sheet']
        except:
            pass
        else:
            save_file_name = f"{self.customer_name}_dashboard_audit_report.xlsx"
            file_path = 'data/'
            current_name = Path(f'{file_path}{save_file_name}')
            

            
            #dectect if previous report is in folder and if so rename file 
            #get list of file in current folder
            dir_list = os.listdir("./data")
            logging.info(f'files in data folder {dir_list}')
            #if report with same name exists rename 
            if save_file_name in dir_list:
                #get date and time, file was created
                ti_c = os.path.getctime(file_path + save_file_name) #time created in seconds
                c_ti = time.ctime(ti_c) #Convert a time expressed in seconds to human readable
                time_in_string = str(c_ti)
                #remove unwanted character from time formate
                for character in "/: ":
                    time_in_string = time_in_string.replace(character,'')

   
                #print(c_ti.format())
                logging.warning(f'previous report file found {save_file_name}, was created {c_ti}')
              
                #check if there's a "backup" folder, if not create one
                if not os.path.exists('./data/backup'):
                    logging.warning("backup folder not present in data folder- creating in data folder")
                    os.mkdir("./data/backup")

                new_file_name = f'{self.customer_name}_BackupReport_{time_in_string}.xlsx'
                new_name = Path(f'{file_path}backup/{new_file_name}')
                current_name.rename(new_name)



            self.audit_report_work_book.save(current_name)


    def day_name(self,day_number:int)-> str:
        logging.debug(f'method - day_name')
        if day_number == 0:
            return 'Monday'
        elif day_number == 1:
            return 'Tuesday'
        elif day_number == 2:
            return 'Wednesday'
        elif day_number == 3:
            return 'Thursday'
        elif day_number == 4:
            return 'Friday'
        elif day_number == 5:
            return 'Saturday'
        elif day_number == 6:
            return 'Sunday'
        else:
            return 'day not know'

    def report_excel_understand_daily_usage(self)->None:
        logging.debug(f'method - report_excel_understand_daily_usage')
        """
        An excel report that will show daily usage over time such as, number of different dashboard used and list of those dashboard for each day. 
        Plus number of unique users who used the dash boards and list of those users. 
        """

        #create excel worksheet

        self.audit_report_work_book.create_sheet("daily_dashboard_usage_overview")
        daily_usage_work_sheet = self.audit_report_work_book["daily_dashboard_usage_overview"]

        #create column tittles
        daily_usage_work_sheet.cell(1,1).value = "week number"
        daily_usage_work_sheet.cell(1,2).value = "date"
        daily_usage_work_sheet.cell(1,3).value = "day"
        daily_usage_work_sheet.cell(1,4).value = "Number of unique dashboards viewed"
        daily_usage_work_sheet.cell(1,5).value = "List of dashboards viewed"
        daily_usage_work_sheet.cell(1,6).value = "Number of unique users that viewed dashboards"
        daily_usage_work_sheet.cell(1,7).value = "List of users that viewed dashboards"

        #output data to worksheet
        row_number = 2
        merged_cells = True
        for w_data in self.week_day_data_structure:
            for d_data in self.week_day_data_structure[w_data]:

                #for week number we those cells will be mergered for each week
                if merged_cells:
                    daily_usage_work_sheet.merge_cells(start_row=row_number, start_column=1,end_row=(row_number+len(self.week_day_data_structure[w_data])-1), end_column=1)
                    daily_usage_work_sheet.cell(row_number,1).value = w_data
                    merged_cells = False
                if d_data['day_number'] == 6:
                    merged_cells = True

                daily_usage_work_sheet.cell(row_number,2).value = d_data['date']
                daily_usage_work_sheet.cell(row_number,3).value = self.day_name(d_data['day_number'])
                daily_usage_work_sheet.cell(row_number,4).value = len(set(d_data['dashboards']))
                daily_usage_work_sheet.cell(row_number,5).value = str([*set(d_data['dashboards'])])
                daily_usage_work_sheet.cell(row_number,6).value = len(set(d_data['users']))
                daily_usage_work_sheet.cell(row_number,7).value = str([*set(d_data['users'])])

                row_number += 1

    def run_standard_usage_reports(self) -> None:
        logging.debug(f'method - run_standard_usage_reports')
        self.report_of_users_using_aternity()
        self.report_what_dashboard_users_used()
        self.report_what_user_used_which_dashboard()


if __name__ == "__main__":
    
    
    aternity_instance_list = []

    #parse in number of day, not required, default it 30 days 
    #parse in file path for customer atternity instance csv file, not required
    #parse in username, required 
    parse = argparse.ArgumentParser()
    parse.add_argument("-u",'--username', type = str, help="username for running rest api report request",required=True)
    parse.add_argument("-d",'--days', type = int, help="number of days to filter on, max and default is 30 days",required=False)
    parse.add_argument("-f",'--file_name', type = str, help="file path for csv file that has  customer atternity instance information",required=False)
    parse.add_argument("-b",'--debug', type = str, help="change the default debug (30) level : 10 - Debug; 20 - Info; 30 - Warning; 40 - Error; 50- Critical\n",required=False)
    
    


    parsed_arg = parse.parse_args()

    username = parsed_arg.username
    passwd = getpass.getpass(f"enter password for user {username}:-")

    #setup logging
    if parsed_arg.debug:        
        if parsed_arg.debug in DEBUG_LEVEL_LIST:
            logging.basicConfig(filename='logs.log', level=int(parsed_arg.debug), filemode='w',format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',datefmt='%m/%d/%Y %H:M: %S')
        else:
            logging.basicConfig(filename='logs.log', level=logging.WARNING, filemode='w',format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',datefmt='%m/%d/%Y %H:M: %S') 
            logging.warning(f'debug change option entered but failed - {parsed_arg.debug}')
    else:
        logging.basicConfig(filename='logs.log', level=logging.INFO, filemode='w',format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',datefmt='%m/%d/%Y %H:M: %S') 


    logging.info(f'username - {username}')
    logging.info(f'password - {passwd[0:4]}*****************{passwd[-4:]}')

    #check if number of days to run report on was entered by user, if not set to 40 days
    if parsed_arg.days:
        number_of_report_days = parsed_arg.days
    else:
        number_of_report_days = 30

    if number_of_report_days > 30:
        print(f"Number of days enter ({number_of_report_days}), must be 30 or less")
        logging.info(f"Quit probgram - Number of days enter ({number_of_report_days}), must be 30 or less")
        quit()

    #check if customer anternity instance file path was entered, if not set to default
    #customer anternity instance file, is a CSV file with list of aternity customers and their coresponding aternity instance web url 
    if parsed_arg.file_name:
        customer_aternity_instance_file_path = parsed_arg.file_name
    else:
        customer_aternity_instance_file_path = DEFAULT_CUSTOMER_ATERNITY_INSTANCE_FILE_PATH

    logging.info(f"customer file information location - {customer_aternity_instance_file_path}")

    #try and open customer aternity instance information file. If failed to open file, set flag "flag_file_opened" to false otherwise to true
    try:
        with open(customer_aternity_instance_file_path,'r') as data_file:
            data = csv.DictReader(data_file)
            customer_aternity_instance_info = list(data)
    except:
        flag_file_opened = False
    else:
        flag_file_opened = True

    logging.info(f'customer file opened = {flag_file_opened}')
    #if flag_file_opened false, script end
    if not flag_file_opened:
        print(f"couldn't open {customer_aternity_instance_file_path}")
        logging.critical(f"couldn't open {customer_aternity_instance_file_path}")
        #script end
    else:

        # check data can be parsed from aternity instance information file, by looking for correct expected dictionary keys "AccountName" and "URL".
        #if not present script end
        if not customer_aternity_instance_info[0].get("AccountName") or not customer_aternity_instance_info[0].get("URL"):
            print("data can not be parsed from aternity instance file {customer_aternity_instance_file_path}, check format. fields/columns names are case sensitive 'AccountName,URL'")
            logging.critical("data can not be parsed from aternity instance file {customer_aternity_instance_file_path}, check format. fields/columns names are case sensitive 'AccountName,URL'")
            #script end
        else:
            #from list, create list of aternity_instance objects for each customer
            for aternity_instance_info in customer_aternity_instance_info:
                customer_aternity_instance = {'aternity_instance':CustomerAternityInstance(aternity_instance_info['AccountName'],aternity_instance_info['URL'],number_of_report_days,username,passwd)}
                aternity_instance_list.append(customer_aternity_instance)
            
            
            #line below shoudl be normally marked out. is used for testing new reporst on old backed up data
            #aternity_instance_list[4]['aternity_instance'].report_excel_understand_daily_usage()
            
        #run rest api report request, getting back a boolean result regarding if running report was sucessful
        
        for aternity_instance in aternity_instance_list:
            aternity_instance["report_status"]=aternity_instance['aternity_instance'].run_rest_api_report_request()
        
        for aternity_instance in aternity_instance_list:
            if aternity_instance["report_status"]:
                aternity_instance['aternity_instance'].excel_audit_report()
               

    